import sqlite3
from sqlite3 import connect
from excel import reader

# Устанавливаем соединение с базой данных
connection = sqlite3.connect('Рассеянный склероз_ДСМ.xlsx')
cursor = connection.cursor()

# Создаем таблицу Users
cursor.execute('''
CREATE TABLE IF NOT EXISTS Users (
id INTEGER PRIMARY KEY,
Тип товара TEXT,
Полное наименование TEXT,
МНН TEXT,
Лекарственная форма TEXT,
Фирма-производитель TEXT,
Корпорация TEXT,
Фармакотерапевтическая группа TEXT,
Год INTEGER,
Месяц INTEGER,
Объем (цена розн.-уп.розн.), рубли INTEGER,
Объем (розница), упак. INTEGER,
Объем (цена опт.-уп.розн.), рубли INTEGER,
)
''')

# Сохраняем изменения и закрываем соединение
connection.commit()
connection.close()




import os
import sqlite3
import openpyxl


def export_to_sqlite():
    '''Экспорт данных из xlsx в sqlite'''

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'auto.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute(('''
CREATE TABLE IF NOT EXISTS Users (
id INTEGER PRIMARY KEY,
Тип товара TEXT,
Полное наименование TEXT,
МНН TEXT,
Лекарственная форма TEXT,
Фирма-производитель TEXT,
Корпорация TEXT,
Фармакотерапевтическая группа TEXT,
Год INTEGER,
Месяц INTEGER,
Объем (цена розн.-уп.розн.), рубли INTEGER,
Объем (розница), упак. INTEGER,
Объем (цена опт.-уп.розн.), рубли INTEGER,
)
'''))

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('Рассеянный склероз_ДСМ.xlsx', data_only=True)
    sheet = file_to_read['Sheet1']

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 5):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

    # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO cars VALUES (null, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (chunk = []
      chunk.append(line[1])
        chunk.append(line[2])
        chunk.append(line[3])
        chunk.append(line[4])
        chunk.append(line[5])
        chunk.append(line[6])
        chunk.append(line[7])
        chunk.append(line[8])
        chunk.append(line[9])
        chunk.append(line[10])
        chunk.append(line[11])
        chunk.append(line[12])
        chunk.append(line[13])
        chunk.append(line[14])
        chunk.append(line[15])
        chunk.append(line[16])
        chunk.append(line[17])
        chunk.append(int(line[18]))
        chunk.append(int(line[19]))
        chunk.append(int(line[20]))
        chunk.append(int(line[21]))
        chunk.append(int(line[22]))
        chunk.append(line[23])
        chunk.append(line[24])
        chunk.append(line[25])
        data.append(chunk))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'auto.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM cars")
    connect.commit()
    connect.close()


# Запуск функции
export_to_sqlite()
